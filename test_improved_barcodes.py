import barcode
from barcode.writer import ImageWriter
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
import os
import sys


def create_barcode_with_padding(code, excel_row_height=85):
    """
    Создает штрих-код с отступами, который заполнит высоту ячейки Excel
    """
    barcode_class = barcode.get_barcode_class('code128')

    # Настройки для улучшенного качества
    writer = ImageWriter()
    writer.set_options({
        'module_height': 20.0,  # Увеличенная высота модуля для лучшей читаемости
        'module_width': 0.55,   # Увеличенная ширина модуля для лучшей читаемости
        'quiet_zone': 1.0,      # Увеличенные поля для лучшей читаемости
        'write_text': False,    # Без текста под штрих-кодом
        'dpi': 300,             # Высокое качество печати
        'text_distance': 5,     # Расстояние текста от штрих-кода (если включено)
        'font_size': 10         # Размер шрифта (если включено)
    })

    barcode_obj = barcode_class(code, writer=writer)
    
    # Создаем изображение
    barcode_img = barcode_obj.render()

    # Получаем размеры
    width, height = barcode_img.size

    # Рассчитываем целевую высоту с учетом отступов
    target_height_px = int(excel_row_height * 3.78)  # Конвертируем пункты в пиксели (примерно)

    # Масштабируем для заполнения ячейки
    scale = target_height_px / height
    new_width = int(width * scale)

    # Изменяем размер с высоким качеством
    scaled_img = barcode_img.resize((new_width, target_height_px), Image.Resampling.LANCZOS)

    # ДОБАВЛЯЕМ ОТСТУПЫ К ИЗОБРАЖЕНИЮ
    # Увеличенные отступы для лучшей читаемости
    top_padding = 25  # Увеличенный отступ сверху
    bottom_padding = 25  # Увеличенный отступ снизу
    left_padding = 30  # Увеличенный отступ слева
    right_padding = 30  # Увеличенный отступ справа

    # Создаем новое изображение с отступами
    padded_width = scaled_img.width + left_padding + right_padding
    padded_height = scaled_img.height + top_padding + bottom_padding

    # Создаем белый фон
    padded_img = Image.new('RGB', (padded_width, padded_height), 'white')

    # Вставляем штрих-код с отступами
    padded_img.paste(scaled_img, (left_padding, top_padding))

    return padded_img


def create_simple_excel_test_with_padding():
    """
    Простой тестовый файл с отступами
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Структура: A - номер, B - код, C - штрих-код
    ws['A1'] = "№"
    ws['B1'] = "Код"

    # Настраиваем ширину колонок
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50

    ROW_HEIGHT = 90  # Увеличенная высота для лучшего качества

    print("Создание тестового файла с отступами и улучшенным качеством...")

    for i in range(1, 6):  # Только 5 для быстрого теста
        code = f"CC{i:03d}"

        # Простые данные
        ws[f'A{i + 1}'] = i
        ws[f'B{i + 1}'] = code

        # Высота строки
        ws.row_dimensions[i + 1].height = ROW_HEIGHT

        try:
            # Создаем штрих-код с отступами
            barcode_img = create_barcode_with_padding(code, ROW_HEIGHT)

            # Сохраняем с высоким качеством
            img_bytes = BytesIO()
            barcode_img.save(img_bytes, format='PNG', dpi=(300, 300))
            img_bytes.seek(0)

            excel_img = ExcelImage(img_bytes)
            excel_img.height = 85
            excel_img.width = 220  # Ширина с отступами

            ws.add_image(excel_img, f'C{i + 1}')

            print(f"  Создан с отступами: {code}")

        except Exception as e:
            print(f"  Пропущен {code}: {e}")
            ws[f'C{i + 1}'] = code
            ws[f'C{i + 1}'].alignment = openpyxl.styles.Alignment(
                horizontal='center',
                vertical='center'
            )

    # Сохраняем
    filename = 'тест_штрихкоды_с_улучшенным_качеством.xlsx'
    wb.save(filename)
    print(f"\n✓ Тестовый файл создан: {filename}")

    return filename


if __name__ == "__main__":
    print("=" * 60)
    print("ТЕСТИРОВАНИЕ УЛУЧШЕННОГО КАЧЕСТВА ШТРИХ-КОДОВ")
    print("=" * 60)
    
    file_path = create_simple_excel_test_with_padding()
    
    print("\n" + "=" * 60)
    print("НАСТРОЙКИ УЛУЧШЕННОГО КАЧЕСТВА:")
    print("1. module_height = 20.0  # Увеличенная высота модуля")
    print("2. module_width = 0.55   # Увеличенная ширина модуля")
    print("3. dpi = 300             # Высокое качество печати")
    print("4. quiet_zone = 1.0      # Увеличенные поля для лучшей читаемости")
    print("5. ROW_HEIGHT = 90       # Увеличенная высота строки")
    print("=" * 60)