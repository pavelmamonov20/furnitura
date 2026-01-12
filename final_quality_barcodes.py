import barcode
from barcode.writer import ImageWriter
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
import os
import sys


def create_barcode_with_padding(code, excel_row_height=85):
    """
    Создает штрих-код с отступами, который заполнит высоту ячейки Excel
    """
    barcode_class = barcode.get_barcode_class('code128')

    # Настройки для улучшенного качества
    writer = ImageWriter()
    writer.set_options({
        'module_height': 25.0,  # Увеличенная высота модуля для лучшей читаемости
        'module_width': 0.7,   # Увеличенная ширина модуля для лучшей читаемости
        'quiet_zone': 1.5,      # Увеличенные поля для лучшей читаемости
        'write_text': False,    # Без текста под штрих-кодом
        'dpi': 600,             # Очень высокое качество для четкости
        'text_distance': 5,     # Расстояние текста от штрих-кода (если включено)
        'font_size': 10         # Размер шрифта (если включено)
    })

    barcode_obj = barcode_class(code, writer=writer)
    
    # Создаем изображение
    barcode_img = barcode_obj.render()

    # Получаем размеры
    width, height = barcode_img.size

    # Рассчитываем целевую высоту с учетом отступов
    target_height_px = int(excel_row_height * 3.78)  # Конвертируем пункты в пиксели (примерно)

    # Масштабируем для заполнения ячейки
    scale = target_height_px / height
    new_width = int(width * scale)

    # Изменяем размер с высоким качеством
    scaled_img = barcode_img.resize((new_width, target_height_px), Image.Resampling.LANCZOS)

    # ДОБАВЛЯЕМ ОТСТУПЫ К ИЗОБРАЖЕНИЮ
    # Увеличенные отступы для лучшей читаемости
    top_padding = 30  # Увеличенный отступ сверху
    bottom_padding = 30  # Увеличенный отступ снизу
    left_padding = 40  # Увеличенный отступ слева
    right_padding = 40  # Увеличенный отступ справа

    # Создаем новое изображение с отступами
    padded_width = scaled_img.width + left_padding + right_padding
    padded_height = scaled_img.height + top_padding + bottom_padding

    # Создаем белый фон
    padded_img = Image.new('RGB', (padded_width, padded_height), 'white')

    # Центрируем штрих-код в области с отступами
    center_x = (padded_width - scaled_img.width) // 2
    center_y = (padded_height - scaled_img.height) // 2
    
    # Вставляем штрих-код по центру
    padded_img.paste(scaled_img, (center_x, center_y))

    return padded_img


def create_simple_excel_test():
    """
    Простой тестовый файл с центровкой и улучшенным качеством
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Структура: A - номер, B - код, C - штрих-код
    ws['A1'] = "№"
    ws['B1'] = "Код"

    # Настраиваем ширину колонок
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60  # Увеличена для центровки

    ROW_HEIGHT = 90  # Увеличенная высота для лучшего качества

    print("Создание тестового файла с центровкой и улучшенным качеством...")

    for i in range(1, 6):  # Только 5 для теста
        code = f"CC{i:03d}"

        # Простые данные
        ws[f'A{i + 1}'] = i
        ws[f'B{i + 1}'] = code

        # Высота строки
        ws.row_dimensions[i + 1].height = ROW_HEIGHT

        try:
            # Создаем штрих-код с центровкой
            barcode_img = create_barcode_with_padding(code, ROW_HEIGHT)

            # Сохраняем с очень высоким качеством
            img_bytes = BytesIO()
            barcode_img.save(img_bytes, format='PNG', dpi=(600, 600))
            img_bytes.seek(0)

            excel_img = ExcelImage(img_bytes)
            excel_img.height = 85
            excel_img.width = 300  # Ширина для центровки

            # Вставляем изображение в Excel
            ws.add_image(excel_img, f'C{i + 1}')
            
            # Центрируем в ячейке
            ws[f'C{i + 1}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            print(f"  Создан с центровкой: {code}")

        except Exception as e:
            print(f"  Пропущен {code}: {e}")
            ws[f'C{i + 1}'] = code
            ws[f'C{i + 1}'].alignment = openpyxl.styles.Alignment(
                horizontal='center',
                vertical='center'
            )

    # Сохраняем
    filename = 'тест_штрихкоды_с_улучшенным_качеством.xlsx'
    wb.save(filename)
    print(f"\n✓ Тестовый файл создан: {filename}")

    return filename


def check_excel_file_for_duplicates(filename):
    """
    Проверяет Excel файл на наличие дубликатов изображений
    """
    if not os.path.exists(filename):
        print(f"Файл {filename} не найден")
        return

    print(f"\nПроверка файла {filename} на дубликаты...")

    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        # Считаем изображения
        image_count = len(ws._images) if hasattr(ws, '_images') else 0
        row_count = ws.max_row

        print(f"  Всего строк: {row_count}")
        print(f"  Всего изображений: {image_count}")

        if image_count == row_count - 1:  # -1 для заголовка
            print("  ✓ Все в порядке: по одному изображению на строку")
        elif image_count > row_count - 1:
            print(f"  ⚠ Возможные дубли: изображений больше чем строк")
        else:
            print(f"  ⚠ Не все строки содержат изображения")

    except Exception as e:
        print(f"  Ошибка при проверке: {e}")


if __name__ == "__main__":
    print("=" * 60)
    print("ГЕНЕРАТОР ШТРИХ-КОДОВ С ЦЕНТРОВКОЙ И УЛУЧШЕННЫМ КАЧЕСТВОМ")
    print("=" * 60)

    # Автоматически запускаем тестовый файл
    print("Создание тестового файла с центровкой и улучшенным качеством...")
    file_path = create_simple_excel_test()
    check_excel_file_for_duplicates(file_path)

    print("\n" + "=" * 60)
    print("НАСТРОЙКИ УЛУЧШЕННОГО КАЧЕСТВА:")
    print("1. module_height = 25.0  # Увеличенная высота модуля")
    print("2. module_width = 0.7    # Увеличенная ширина модуля")
    print("3. dpi = 600             # Очень высокое качество печати")
    print("4. quiet_zone = 1.5      # Увеличенные поля для лучшей читаемости")
    print("5. ROW_HEIGHT = 90       # Увеличенная высота строки")
    print("6. Центрирование штрих-кода в изображении и в ячейке Excel")
    print("=" * 60)