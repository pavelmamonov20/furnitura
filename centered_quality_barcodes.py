import barcode
from barcode.writer import ImageWriter
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
import os
import sys


def create_barcode_with_padding(code, excel_row_height=85):
    """
    Создает штрих-код с отступами, который заполнит высоту ячейки Excel
    """
    barcode_class = barcode.get_barcode_class('code128')

    # Настройки для улучшенного качества
    writer = ImageWriter()
    writer.set_options({
        'module_height': 25.0,  # Еще больше увеличена высота модуля для лучшей читаемости
        'module_width': 0.7,   # Еще больше увеличена ширина модуля для лучшей читаемости
        'quiet_zone': 1.5,      # Увеличенные поля для лучшей читаемости
        'write_text': False,    # Без текста под штрих-кодом
        'dpi': 600,             # Очень высокое качество для четкости
        'text_distance': 5,     # Расстояние текста от штрих-кода (если включено)
        'font_size': 10         # Размер шрифта (если включено)
    })

    barcode_obj = barcode_class(code, writer=writer)
    
    # Создаем изображение
    barcode_img = barcode_obj.render()

    # Получаем размеры
    width, height = barcode_img.size

    # Рассчитываем целевую высоту с учетом отступов
    target_height_px = int(excel_row_height * 3.78)  # Конвертируем пункты в пиксели (примерно)

    # Масштабируем для заполнения ячейки
    scale = target_height_px / height
    new_width = int(width * scale)

    # Изменяем размер с высоким качеством
    scaled_img = barcode_img.resize((new_width, target_height_px), Image.Resampling.LANCZOS)

    # ДОБАВЛЯЕМ ОТСТУПЫ К ИЗОБРАЖЕНИЮ
    # Увеличенные отступы для лучшей читаемости
    top_padding = 30  # Увеличенный отступ сверху
    bottom_padding = 30  # Увеличенный отступ снизу
    left_padding = 40  # Увеличенный отступ слева
    right_padding = 40  # Увеличенный отступ справа

    # Создаем новое изображение с отступами
    padded_width = scaled_img.width + left_padding + right_padding
    padded_height = scaled_img.height + top_padding + bottom_padding

    # Создаем белый фон
    padded_img = Image.new('RGB', (padded_width, padded_height), 'white')

    # Центрируем штрих-код в области с отступами
    center_x = (padded_width - scaled_img.width) // 2
    center_y = (padded_height - scaled_img.height) // 2
    
    # Вставляем штрих-код по центру
    padded_img.paste(scaled_img, (center_x, center_y))

    return padded_img


def create_excel_with_centered_barcodes(start=1, end=200, output_file='штрихкоды_с_центровкой.xlsx'):
    """
    Создает Excel файл с правильно размещенными штрих-кодами с центровкой
    """
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Штрих-коды"

    # Устанавливаем заголовки
    ws['A1'] = "№"
    ws['B1'] = "Код текстом"

    # Настраиваем стиль
    header_font = openpyxl.styles.Font(bold=True, size=11)
    header_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    for cell in ['A1', 'B1']:
        ws[cell].font = header_font
        ws[cell].alignment = header_alignment

    # Ширина колонок (увеличили колонку C для центровки)
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60  # Ещё больше увеличили для центровки

    # Высота строки 30 мм = 90 пунктов
    ROW_HEIGHT = 90  # Увеличили для лучшего качества

    print(f"Создание Excel файла со штрих-кодами с центровкой...")
    print(f"Высота строки: {ROW_HEIGHT} пунктов (30 мм)")
    print(f"Отступы: 40px слева/справа, 30px сверху/снизу")
    print("=" * 60)

    for i in range(start, end + 1):
        row_num = i + 1  # Строка 1 - заголовок, данные с 2 строки
        code = f"CC{i:03d}"

        # Записываем данные
        ws[f'A{row_num}'] = i
        ws[f'B{row_num}'] = code

        # Выравнивание
        ws[f'A{row_num}'].alignment = header_alignment
        ws[f'B{row_num}'].alignment = header_alignment

        # Добавляем границы
        for col in ['A', 'B']:
            ws[f'{col}{row_num}'].border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )

        # Создаем и вставляем штрих-код с центровкой
        try:
            print(f"  Создаю с центровкой: {code}")

            barcode_img = create_barcode_with_padding(code, ROW_HEIGHT)

            # Сохраняем в байтовый поток с высоким качеством
            img_bytes = BytesIO()
            barcode_img.save(img_bytes, format='PNG', dpi=(600, 600))  # Очень высокое качество
            img_bytes.seek(0)

            # Вставляем в Excel
            img = ExcelImage(img_bytes)

            # Устанавливаем размеры в Excel
            img.height = 85  # Высота ячейки (30 мм)
            img.width = 300  # Фиксированная ширина для центровки

            # Вставляем в колонку C с центровкой
            ws.add_image(img, f'C{row_num}')
            
            # Центрируем изображение в ячейке путем настройки ширины колонки и выравнивания
            ws[f'C{row_num}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        except Exception as e:
            print(f"  Ошибка: {e}")
            # В случае ошибки записываем код текстом
            ws[f'C{row_num}'] = code
            ws[f'C{row_num}'].alignment = header_alignment
            continue

        # Устанавливаем высоту строки
        ws.row_dimensions[row_num].height = ROW_HEIGHT

        # Прогресс
        if i % 20 == 0:
            print(f"  Прогресс: {i} из {end}")

    # Сохраняем файл
    print("\nСохранение файла...")

    # Пробуем сохранить в разные места
    save_locations = [
        output_file,
        os.path.join(os.path.expanduser('~'), 'Desktop', output_file),
        os.path.join(os.path.expanduser('~'), 'Documents', output_file)
    ]

    saved = False
    final_path = ""

    for location in save_locations:
        try:
            wb.save(location)
            print(f"✓ Файл сохранен: {location}")
            saved = True
            final_path = location
            break
        except Exception as e:
            print(f"  Не удалось сохранить в {location}: {e}")
            continue

    if not saved:
        # Последняя попытка с другим именем
        import time
        backup_name = f'штрихкоды_{int(time.time())}.xlsx'
        wb.save(backup_name)
        print(f"✓ Файл сохранен как: {backup_name}")
        final_path = backup_name

    # Выводим информацию о файле
    if os.path.exists(final_path):
        file_size = os.path.getsize(final_path) / 1024 / 1024
        print(f"\n✓ Размер файла: {file_size:.2f} MB")
        print(f"✓ Всего строк: {end - start + 1}")
        print(f"✓ Высота каждой строки: {ROW_HEIGHT} пунктов (30 мм)")
        print(f"✓ Штрих-коды центрированы и имеют улучшенное качество")

    return final_path


def create_simple_excel_test_with_centering():
    """
    Простой тестовый файл с центровкой и улучшенным качеством
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Структура: A - номер, B - код, C - штрих-код
    ws['A1'] = "№"
    ws['B1'] = "Код"

    # Настраиваем ширину колонок
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60  # Увеличена для центровки

    ROW_HEIGHT = 90  # Увеличенная высота для лучшего качества

    print("Создание тестового файла с центровкой и улучшенным качеством...")

    for i in range(1, 6):  # Только 5 для теста
        code = f"CC{i:03d}"

        # Простые данные
        ws[f'A{i + 1}'] = i
        ws[f'B{i + 1}'] = code

        # Высота строки
        ws.row_dimensions[i + 1].height = ROW_HEIGHT

        try:
            # Создаем штрих-код с центровкой
            barcode_img = create_barcode_with_padding(code, ROW_HEIGHT)

            # Сохраняем с очень высоким качеством
            img_bytes = BytesIO()
            barcode_img.save(img_bytes, format='PNG', dpi=(600, 600))
            img_bytes.seek(0)

            excel_img = ExcelImage(img_bytes)
            excel_img.height = 85
            excel_img.width = 300  # Ширина для центровки

            # Вставляем изображение в Excel
            ws.add_image(excel_img, f'C{i + 1}')
            
            # Центрируем в ячейке
            ws[f'C{i + 1}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            print(f"  Создан с центровкой: {code}")

        except Exception as e:
            print(f"  Пропущен {code}: {e}")
            ws[f'C{i + 1}'] = code
            ws[f'C{i + 1}'].alignment = openpyxl.styles.Alignment(
                horizontal='center',
                vertical='center'
            )

    # Сохраняем
    filename = 'тест_штрихкоды_с_улучшенным_качеством.xlsx'
    wb.save(filename)
    print(f"\n✓ Тестовый файл создан: {filename}")

    return filename


def check_excel_file_for_duplicates(filename):
    """
    Проверяет Excel файл на наличие дубликатов изображений
    """
    if not os.path.exists(filename):
        print(f"Файл {filename} не найден")
        return

    print(f"\nПроверка файла {filename} на дубликаты...")

    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        # Считаем изображения
        image_count = len(ws._images) if hasattr(ws, '_images') else 0
        row_count = ws.max_row

        print(f"  Всего строк: {row_count}")
        print(f"  Всего изображений: {image_count}")

        if image_count == row_count - 1:  # -1 для заголовка
            print("  ✓ Все в порядке: по одному изображению на строку")
        elif image_count > row_count - 1:
            print(f"  ⚠ Возможные дубли: изображений больше чем строк")
        else:
            print(f"  ⚠ Не все строки содержат изображения")

    except Exception as e:
        print(f"  Ошибка при проверке: {e}")


if __name__ == "__main__":
    print("=" * 60)
    print("ГЕНЕРАТОР ШТРИХ-КОДОВ С ЦЕНТРОВКОЙ И УЛУЧШЕННЫМ КАЧЕСТВОМ")
    print("=" * 60)

    # Автоматически запускаем тестовый файл (вариант 2)
    print("Создание тестового файла с центровкой и улучшенным качеством...")
    file_path = create_simple_excel_test_with_centering()
    check_excel_file_for_duplicates(file_path)

    print("\n" + "=" * 60)
    print("НАСТРОЙКИ УЛУЧШЕННОГО КАЧЕСТВА:")
    print("1. module_height = 25.0  # Увеличенная высота модуля")
    print("2. module_width = 0.7    # Увеличенная ширина модуля")
    print("3. dpi = 600             # Очень высокое качество печати")
    print("4. quiet_zone = 1.5      # Увеличенные поля для лучшей читаемости")
    print("5. ROW_HEIGHT = 90       # Увеличенная высота строки")
    print("6. Центрирование штрих-кода в изображении и в ячейке Excel")
    print("=" * 60)