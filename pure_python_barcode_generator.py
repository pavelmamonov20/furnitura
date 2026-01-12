"""
Pure Python Vector-Based Barcode Generator for Excel
Creates high-quality barcodes using SVG manipulation and converts them to PNG images
for optimal clarity and perfect centering in Excel cells.
"""

import barcode
from barcode.writer import SVGWriter
from xml.etree import ElementTree as ET
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, Border, Side
import os
import sys


def create_svg_barcode(code):
    """
    Creates an SVG barcode using python-barcode library
    """
    # Get the CODE128 barcode class
    barcode_class = barcode.get_barcode_class('code128')
    
    # Create SVG writer with custom settings for better quality
    writer = SVGWriter()
    writer.set_options({
        'module_height': 20.0,      # Increased height for better readability
        'module_width': 0.6,        # Slightly wider bars for clarity
        'quiet_zone': 15.0,         # More space on both sides
        'font_size': 12,            # Larger human-readable text
        'text_distance': 7,         # Distance between barcode and text
        'background': 'white',      # White background
        'foreground': 'black',      # Black bars
        'write_text': True          # Include human-readable text
    })
    
    # Create the barcode object
    barcode_obj = barcode_class(code, writer=writer)
    
    # Render to SVG string
    svg_string = barcode_obj.render()
    
    return svg_string


def parse_svg_and_create_png(svg_content, target_width=300, target_height=150):
    """
    Parses SVG content and creates a high-quality PNG using PIL
    This function manually parses the SVG and draws it with PIL
    """
    from PIL import Image, ImageDraw
    
    # Parse the SVG content
    root = ET.fromstring(svg_content)
    
    # Extract dimensions from the SVG
    width_attr = root.get('width', '200')
    height_attr = root.get('height', '100')
    
    # Remove units if present (like 'mm', 'px', '%') and handle percentages
    import re
    def extract_numeric_value(attr):
        # Handle percentage values by returning a reasonable default
        if '%' in attr:
            # For percentage values, return a standard size based on content
            return 200  # Standard width for barcodes
        else:
            # Remove units like 'mm', 'px' and keep only numeric part
            return float(re.sub(r'[^\d.]', '', attr))
    
    width_val = extract_numeric_value(width_attr)
    height_val = extract_numeric_value(height_attr)
    
    # Create a larger image for high resolution
    scale_factor = 3  # Scale up for better quality
    img_width = int(target_width * scale_factor)
    img_height = int(target_height * scale_factor)
    
    # Calculate scaling factors based on original SVG size
    x_scale = img_width / width_val
    y_scale = img_height / height_val
    
    # Create a white background image
    img = Image.new('RGB', (img_width, img_height), 'white')
    draw = ImageDraw.Draw(img)
    
    # Find the barcode group element
    for elem in root.iter():
        tag = elem.tag.split('}')[-1]  # Remove namespace
        
        if tag == 'rect':
            # Get rectangle attributes
            x = elem.get('x', '0')
            y = elem.get('y', '0')
            rect_width = elem.get('width', '1')
            rect_height = elem.get('height', '10')
            
            # Handle percentage values for individual elements too
            def clean_value(val):
                if '%' in val:
                    # Convert percentage to actual dimension based on assumed original size
                    pct = float(re.sub(r'[^\d.]', '', val)) / 100
                    if 'x' in elem.keys() or 'y' in elem.keys():
                        return pct * width_val
                    else:
                        return pct * height_val
                else:
                    return float(re.sub(r'[^\d.]', '', val))
            
            try:
                x = clean_value(x)
                y = clean_value(y)
                rect_width = clean_value(rect_width)
                rect_height = clean_value(rect_height)
                
                # Scale coordinates and dimensions
                scaled_x = int(x * x_scale)
                scaled_y = int(y * y_scale)
                scaled_width = max(1, int(rect_width * x_scale))  # Ensure minimum width of 1 pixel
                scaled_height = int(rect_height * y_scale)
                
                # Draw the rectangle (bar)
                draw.rectangle([scaled_x, scaled_y, scaled_x + scaled_width, scaled_y + scaled_height], fill='black')
            except ValueError as e:
                # Skip elements that can't be processed
                continue
        
        elif tag == 'text':
            # Handle text elements if needed (optional)
            pass
    
    # Resize back to target dimensions for anti-aliasing effect
    img = img.resize((target_width, target_height), Image.Resampling.LANCZOS)
    
    # Save to bytes
    img_bytes = BytesIO()
    img.save(img_bytes, format='PNG', dpi=(300, 300))  # High DPI for quality
    img_bytes.seek(0)
    
    return img_bytes.getvalue()


def create_high_quality_barcode_image(code, target_width=300, target_height=150):
    """
    Creates a high-quality barcode image using SVG as base and PIL for rendering
    """
    # Generate SVG barcode
    svg_content = create_svg_barcode(code)
    
    # Convert to PNG using PIL
    png_data = parse_svg_and_create_png(svg_content, target_width, target_height)
    
    return png_data


def create_excel_with_centered_barcodes(start=1, end=20, output_file='тест_штрихкоды_с_улучшенным_качеством.xlsx'):
    """
    Creates an Excel file with centered, high-quality barcodes
    """
    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Штрих-коды"

    # Set headers
    ws['A1'] = "№"
    ws['B1'] = "Код текстом"

    # Define styles
    header_font = Font(bold=True, size=11)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply header styles
    for cell in ['A1', 'B1']:
        ws[cell].font = header_font
        ws[cell].alignment = center_alignment
        ws[cell].border = thin_border

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50  # Wider for barcode display

    # Row height for 30mm (approximately 85 points)
    ROW_HEIGHT = 90

    print(f"Creating Excel file with high-quality centered barcodes...")
    print(f"Row height: {ROW_HEIGHT} points (30 mm)")
    print("=" * 60)

    for i in range(start, end + 1):
        row_num = i + 1  # Data starts from row 2 (header is row 1)
        code = f"CC{i:03d}"

        # Write data
        ws[f'A{row_num}'] = i
        ws[f'B{row_num}'] = code

        # Apply styles to data cells
        for col in ['A', 'B']:
            ws[f'{col}{row_num}'].alignment = center_alignment
            ws[f'{col}{row_num}'].border = thin_border

        # Create and insert the barcode image
        try:
            print(f"  Creating high-quality barcode: {code}")
            
            # Create the barcode image
            barcode_img_data = create_high_quality_barcode_image(code, 300, 150)

            # Create Excel image from bytes
            img_bytes = BytesIO(barcode_img_data)
            excel_img = ExcelImage(img_bytes)

            # Set image properties for centering
            excel_img.width = 250  # Adjust as needed
            excel_img.height = 80  # Adjust as needed

            # Add image to worksheet in column C
            ws.add_image(excel_img, f'C{row_num}')
            
            # Center the image by setting cell alignment
            ws[f'C{row_num}'].alignment = center_alignment

        except Exception as e:
            print(f"  Error creating barcode for {code}: {e}")
            # Fallback: write code as text
            ws[f'C{row_num}'] = code
            ws[f'C{row_num}'].alignment = center_alignment
            ws[f'C{row_num}'].border = thin_border
            continue

        # Set row height
        ws.row_dimensions[row_num].height = ROW_HEIGHT

        # Progress indicator
        if i % 5 == 0:
            print(f"  Progress: {i} out of {end}")

    # Save the workbook
    print("\nSaving file...")
    
    try:
        wb.save(output_file)
        print(f"✓ File saved: {output_file}")
        return output_file
    except PermissionError:
        print(f"Permission denied. Trying alternative location...")
        alternative_path = os.path.join(os.path.expanduser('~'), 'Desktop', output_file)
        wb.save(alternative_path)
        print(f"✓ File saved: {alternative_path}")
        return alternative_path
    except Exception as e:
        print(f"Error saving file: {e}")
        return None


def main():
    """
    Main function to run the pure Python vector-based barcode generator
    """
    print("=" * 60)
    print("PURE PYTHON VECTOR-BASED BARCODE GENERATOR")
    print("=" * 60)
    print("This tool creates high-quality vector-based barcodes using pure Python")
    print("libraries for optimal clarity and perfect centering in Excel.")
    print()

    # Ask for mode
    print("Select mode:")
    print("1. Create test file (20 barcodes)")
    print("2. Create full file (200 barcodes)")
    
    try:
        choice = input("\nEnter your choice (1-2): ").strip()
        
        if choice == "1":
            print("\nCreating test file with 20 high-quality centered barcodes...")
            file_path = create_excel_with_centered_barcodes(1, 20, 'тест_штрихкоды_с_улучшенным_качеством.xlsx')
        elif choice == "2":
            print("\nCreating full file with 200 high-quality centered barcodes...")
            file_path = create_excel_with_centered_barcodes(1, 200, 'штрихкоды_с_улучшенным_качеством.xlsx')
        else:
            print("\nInvalid choice. Creating test file...")
            file_path = create_excel_with_centered_barcodes(1, 20, 'тест_штрихкоды_с_улучшенным_качеством.xlsx')
        
        if file_path:
            file_size = os.path.getsize(file_path) / 1024 / 1024
            print(f"\n✓ File created successfully!")
            print(f"✓ File size: {file_size:.2f} MB")
            print(f"✓ Number of rows: {21 if 'тест_' in file_path else 201} (including header)")
            print(f"✓ Each row has a high-quality centered barcode")
        else:
            print("\n✗ Failed to create file")
            
    except KeyboardInterrupt:
        print("\nOperation cancelled by user.")
    except Exception as e:
        print(f"\nError occurred: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()